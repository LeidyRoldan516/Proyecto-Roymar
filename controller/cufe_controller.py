"""
controller/cufe_controller.py
Orquesta el flujo completo: recibe la lista de CUFEs del frontend,
llama al modelo para consultarlos, y devuelve los resultados.
No sabe nada de HTTP ni de HTML.
"""
import io
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from model.cufe import CufeResultado
from model.dian_client import consultar_cufe
from model.excel_parser import extraer_cufes, ExcelParserError


# ── Configuración ─────────────────────────────────────────────────────────────

MAX_WORKERS = 3       # peticiones paralelas a la DIAN
MAX_CUFES   = 5_000   # límite de seguridad


# ── Entrada: parseo del Excel ─────────────────────────────────────────────────

def procesar_archivo_entrada(ruta: str) -> dict:
    """
    Lee el Excel de entrada y devuelve la metadata + lista de CUFEs.
    Propaga ExcelParserError si el archivo no es válido.
    """
    resultado = extraer_cufes(ruta)
    if len(resultado["cufes"]) > MAX_CUFES:
        raise ExcelParserError(
            f"El archivo contiene {len(resultado['cufes'])} CUFEs. "
            f"El límite por consulta es {MAX_CUFES}."
        )
    return resultado


# ── Consulta masiva ───────────────────────────────────────────────────────────

def consultar_lista(cufes: list[str]) -> list[dict]:
    """
    Consulta la DIAN para cada CUFE en paralelo (MAX_WORKERS hilos).
    Devuelve lista de dicts listos para serializar a JSON, en el mismo orden
    en que llegaron los CUFEs.
    """
    resultados: dict[str, CufeResultado] = {}

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futuros = {executor.submit(consultar_cufe, cufe): cufe for cufe in cufes}
        for futuro in as_completed(futuros):
            cufe = futuros[futuro]
            try:
                resultados[cufe] = futuro.result()
            except Exception as e:
                r = CufeResultado(cufe=cufe, error=str(e), estado=f"Error: {e}")
                resultados[cufe] = r

    # Preservar orden original
    return [_serializar(resultados[c]) for c in cufes]


# ── Generación del Excel de salida ────────────────────────────────────────────

def generar_excel(resultados: list[dict]) -> bytes:
    """
    Genera el Excel de salida en memoria y devuelve los bytes.
    El caller (route) se encarga de enviarlo al browser.
    """
    wb = openpyxl.Workbook()

    _construir_hoja_datos(wb, resultados)
    _construir_hoja_resumen(wb, resultados)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ── Helpers internos ──────────────────────────────────────────────────────────

def _serializar(r: CufeResultado) -> dict:
    return {
        "cufe":           r.cufe,
        "tipo_documento": r.tipo_documento,
        "folio":          r.folio,
        "prefijo":        r.prefijo,
        "divisa":         r.divisa,
        "forma_pago":     r.forma_pago,
        "medio_pago":     r.medio_pago,
        "fecha_emision":  r.fecha_emision,
        "fecha_recepcion":r.fecha_recepcion,
        "nit_emisor":     r.nit_emisor,
        "nombre_emisor":  r.nombre_emisor,
        "nit_receptor":   r.nit_receptor,
        "nombre_receptor":r.nombre_receptor,
        "iva":            r.iva,
        "ica":            r.ica,
        "ic":             r.ic,
        "inc":            r.inc,
        "timbre":         r.timbre,
        "inc_bolsas":     r.inc_bolsas,
        "in_carbono":     r.in_carbono,
        "in_combustibles":r.in_combustibles,
        "ic_datos":       r.ic_datos,
        "icl":            r.icl,
        "inpp":           r.inpp,
        "ibua":           r.ibua,
        "icui":           r.icui,
        "rete_iva":       r.rete_iva,
        "rete_renta":     r.rete_renta,
        "rete_ica":       r.rete_ica,
        "total":          r.total,
        "estado":         r.estado,
        "grupo":          r.grupo,
        "tiene_datos":    r.tiene_datos,
        "error":          r.error,
    }


def _construir_hoja_datos(wb: openpyxl.Workbook, resultados: list[dict]) -> None:
    ws = wb.active
    fecha_hoy = datetime.now().strftime("%Y%m%d")
    ws.title = f"Reporte_DIAN_{fecha_hoy}"

    cabeceras = CufeResultado.cabeceras_excel()
    ws.append(cabeceras)

    # Estilos de cabecera
    fill_header  = PatternFill("solid", fgColor="1E3A5F")
    font_header  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col, _ in enumerate(cabeceras, start=1):
        cell = ws.cell(row=1, column=col)
        cell.fill   = fill_header
        cell.font   = font_header
        cell.alignment = align_center

    # Filas de datos
    for r in resultados:
        ws.append([
            r["tipo_documento"], r["cufe"], r["folio"], r["prefijo"],
            r["divisa"], r["forma_pago"], r["medio_pago"],
            r["fecha_emision"], r["fecha_recepcion"],
            r["nit_emisor"], r["nombre_emisor"],
            r["nit_receptor"], r["nombre_receptor"],
            r["iva"], r["ica"], r["ic"], r["inc"], r["timbre"],
            r["inc_bolsas"], r["in_carbono"], r["in_combustibles"],
            r["ic_datos"], r["icl"], r["inpp"], r["ibua"], r["icui"],
            r["rete_iva"], r["rete_renta"], r["rete_ica"],
            r["total"], r["estado"], r["grupo"],
        ])

    # Anchos de columna
    anchos = [22,64,12,10,8,14,14,18,22,14,30,14,30,
              10,8,8,8,8,10,10,14,10,8,8,8,8,10,12,10,14,30,10]
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho

    ws.freeze_panes = "A2"


def _construir_hoja_resumen(wb: openpyxl.Workbook, resultados: list[dict]) -> None:
    ws = wb.create_sheet("Resumen")
    total   = len(resultados)
    exitosos = sum(1 for r in resultados if r["tiene_datos"])
    errores  = sum(1 for r in resultados if r["error"])
    sin_datos = total - exitosos - errores

    filas = [
        ["RESUMEN DE CONSULTA DIAN", ""],
        ["Fecha de generación", datetime.now().strftime("%d/%m/%Y %H:%M:%S")],
        ["Total CUFEs consultados", total],
        ["Exitosos (con datos)", exitosos],
        ["Sin datos estructurados", sin_datos],
        ["Con error de conexión", errores],
        ["", ""],
        ["Generado por", "Consulta CUFE DIAN v1.0"],
    ]
    for fila in filas:
        ws.append(fila)

    ws["A1"].font = Font(bold=True, size=12, name="Arial")
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 28
